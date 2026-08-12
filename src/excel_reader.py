from openpyxl import load_workbook

def read_part_numbers(file_path):

    workbook = load_workbook(file_path, data_only=True)

    sheet = workbook.active

    part_numbers = []

#! Begin to Read from Row 2
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=1)   #for cell in sheet["A"][1:]:
        
        if cell.value is not None:
            value = str(cell.value).strip()

            if value != "":
                part_numbers.append(value)

    workbook.close()

    return part_numbers
